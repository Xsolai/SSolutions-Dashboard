from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database.models.models import (
    WorkflowReportGuruKF, User, EmailData, QueueStatistics, 
    AllQueueStatisticsData, SoftBookingKF, GuruCallReason, BookingData
)
from app.database.db.db_connection import get_db
from datetime import datetime, timedelta, date
from sqlalchemy import func, or_
from app.database.scehmas import schemas
from app.database.auth import oauth2
from app.src.utils import (
    time_format, time_formatter, validate_user_and_date_permissions_export,
    domains_checker, domains_checker_email, domains_checker_booking
)
from app.src.utils_booking import validate_user_and_date_permissions_booking_export
from typing import Optional, List
import pandas as pd
from io import BytesIO
import openpyxl, io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from app.database.db.db_connection import  get_db, SessionLocal

router = APIRouter(tags=["Export API"])


def get_export_data(call_query, all_call_query, email_query, all_email_query, booking_query, 
                    start_date, end_date, start_date_booking, end_date_booking, domain, company):
    if "guru" in company:
        db = SessionLocal()
        all_calls_data = db.query(
            GuruCallReason.date.label("date"),
            func.sum(GuruCallReason.total_calls).label("all_calls")).filter(
            GuruCallReason.date.between(start_date, end_date)).group_by(GuruCallReason.date).all()
        all_calls = {
            row.date.strftime("%Y-%m-%d") if row.date else None: row.all_calls
            for row in all_calls_data
        }
        # organisch conversion
        # guru_sales = db.query(func.sum(GuruCallReason.guru_sales)).filter(
        #     GuruCallReason.date.between(start_date, end_date)).scalar() or 0
        guru_sales_data = db.query(
            GuruCallReason.date.label("date"),
            func.sum(GuruCallReason.guru_sales).label("guru_sales")).filter(
            GuruCallReason.date.between(start_date, end_date)).group_by(GuruCallReason.date).all()
        guru_sales = {
            row.date.strftime("%Y-%m-%d") if row.date else None: row.guru_sales
            for row in guru_sales_data
        }
        # guru_wrong = db.query(func.sum(GuruCallReason.guru_wrong)).filter(
        #     GuruCallReason.date.between(start_date, end_date)).scalar() or 0
        guru_wrong_data = db.query(
            GuruCallReason.date.label("date"),
            func.sum(GuruCallReason.guru_wrong).label("guru_wrong")).filter(
            GuruCallReason.date.between(start_date, end_date)).group_by(GuruCallReason.date).all()
        guru_wrong = {
            row.date.strftime("%Y-%m-%d") if row.date else None: row.guru_wrong
            for row in guru_wrong_data
        }
        # reason_booking = db.query(func.sum(GuruCallReason.guru_cb_booking)).filter(
        #     GuruCallReason.date.between(start_date, end_date)).scalar() or 0
        reason_booking_data = db.query(
            GuruCallReason.date.label("date"),
            func.sum(GuruCallReason.guru_cb_booking).label("reason_booking")).filter(
            GuruCallReason.date.between(start_date, end_date)).group_by(GuruCallReason.date).all()
        reason_booking = {
            row.date.strftime("%Y-%m-%d") if row.date else None: row.reason_booking
            for row in reason_booking_data
        }
        # guru_bookings = db.query(func.count(BookingData.crs_original_booking_number)).filter(
        #     BookingData.date.between(start_date, end_date),
        #     BookingData.order_agent.notlike("%CB%"),
        #     BookingData.order_agent.like(f"%002%"),
        #     BookingData.crs_status.in_(["OK", "RF"])
        #     ).scalar() or 0
        guru_bookings_data = db.query(
            BookingData.date.label("date"),
            func.count(BookingData.crs_original_booking_number).label("guru_bookings")).filter(
            BookingData.date.between(start_date, end_date),
            BookingData.order_agent.notlike("%CB%"),
            BookingData.order_agent.like(f"%002%"),
            BookingData.crs_status.in_(["OK", "RF"])).group_by(BookingData.date).all()
        guru_bookings = {
            row.date.strftime("%Y-%m-%d") if row.date else None: row.guru_bookings
            for row in guru_bookings_data
        }
        
        # cb conversion
        # cb_sales = db.query(func.sum(GuruCallReason.cb_sales)).filter(
        #     GuruCallReason.date.between(start_date, end_date)).scalar() or 0
        cb_sales_data = db.query(
            GuruCallReason.date.label("date"),
            func.sum(GuruCallReason.cb_sales).label("cb_sales")).filter(
            GuruCallReason.date.between(start_date, end_date)).group_by(GuruCallReason.date).all()
        cb_sales = {
            row.date.strftime("%Y-%m-%d") if row.date else None: row.cb_sales
            for row in cb_sales_data
        }
        # cb_wrong = db.query(func.sum(GuruCallReason.cb_wrong_call)).filter(
        #     GuruCallReason.date.between(start_date, end_date)).scalar() or 0
        cb_wrong_data = db.query(
            GuruCallReason.date.label("date"),
            func.sum(GuruCallReason.cb_wrong_call).label("cb_wrong")).filter(
            GuruCallReason.date.between(start_date, end_date)).group_by(GuruCallReason.date).all()
        cb_wrong = {
            row.date.strftime("%Y-%m-%d") if row.date else None: row.cb_wrong
            for row in cb_wrong_data
        }
        # cb_bookings = db.query(func.count(BookingData.crs_original_booking_number)).filter(
        #     BookingData.date.between(start_date, end_date),
        #     BookingData.order_agent.like("%CB%"),
        #     BookingData.order_agent.like(f"%002%"),
        #     BookingData.crs_status.in_(["OK", "RF"])
        #     ).scalar() or 0
        cb_bookings_data = db.query(
            BookingData.date.label("date"),
            func.count(BookingData.crs_original_booking_number).label("cb_bookings")).filter(
            BookingData.date.between(start_date, end_date),
            BookingData.order_agent.like("%CB%"),
            BookingData.order_agent.like(f"%002%"),
            BookingData.crs_status.in_(["OK", "RF"])).group_by(BookingData.date).all()
        cb_bookings = {
            row.date.strftime("%Y-%m-%d") if row.date else None: row.cb_bookings
            for row in cb_bookings_data
        }
        # print("all_calls:", all_calls)
        # print("guru_sales:", guru_sales)
        # print("guru_wrong:", guru_wrong)
        # print("reason_booking:", reason_booking)
        # print("guru_bookings:", guru_bookings)
        # print("cb_sales:", cb_sales)
        # print("cb_wrong:", cb_wrong)
        # print("cb_bookings:", cb_bookings)
        # Calculate sales conversion rate by date
        organisch_conversion = {}
        cb_conversion = {}
        guru_wrong_call = {}
        cb_wrong_call = {}
        true_guru_calls = {}
        true_cb_calls = {}
        for date in set(all_calls.keys()) | set(guru_sales.keys()) | set(guru_wrong.keys()) | set(reason_booking.keys()) | set(guru_bookings.keys()) | set(cb_sales.keys()) | set(cb_wrong.keys()) | set(cb_bookings.keys()):
            calls = all_calls.get(date, 0)
            guru_sales_val = guru_sales.get(date, 0)
            guru_wrong_val = guru_wrong.get(date, 0)
            reason_booking_val = reason_booking.get(date, 0)
            guru_bookings_val = guru_bookings.get(date, 0)
            cb_sales_val = cb_sales.get(date, 0)
            cb_wrong_val = cb_wrong.get(date, 0)
            cb_bookings_val = cb_bookings.get(date, 0)
            # print(guru_sales, guru_wrong_val, reason_booking, guru_bookings_val)
            guru_wrong_call[date] = guru_sales_val - guru_wrong_val - reason_booking_val
            cb_wrong_call[date] = (cb_sales_val + reason_booking_val) - cb_wrong_val
            
            # Now use dictionary values for true_guru_calls and true_cb_calls
            true_guru_calls[date] = calls - guru_wrong_call[date]
            true_cb_calls[date] = calls - cb_wrong_call[date]
            
            # Use dictionary values for conversion calculations
            organisch_conversion[date] = f"{round(guru_bookings_val/(true_guru_calls[date] if true_guru_calls[date]>0 else 1)*100,4)}%"
            cb_conversion[date] = f"{round(cb_bookings_val/(true_cb_calls[date] if true_cb_calls[date]>0 else 1)*100,4)}%"
            # sales_conversion[date] = f"{round(sales_effective_calls / bookings_success if bookings_success > 0 else 1, 2)}%"
    else:
        all_calls ={}
        guru_wrong_call = {}
        cb_wrong_call = {}
        guru_bookings = {}
        cb_bookings = {}
        organisch_conversion = {}
        cb_conversion = {}
        true_guru_calls = {}
        true_cb_calls = {}
        
    
    yesterday_date = (datetime.now() - timedelta(days=1)).date()
    # print(yesterday_date, yesterday_date_start, yesterday_date_end)
    # Call metrics
    calls = all_call_query.with_entities(func.sum(AllQueueStatisticsData.calls)).filter(
        AllQueueStatisticsData.date.between(start_date, end_date)
    ).scalar() or 0
    calls_sales = call_query.with_entities(func.sum(QueueStatistics.calls)).filter(
        QueueStatistics.date.between(start_date, end_date),
        QueueStatistics.queue_name.like(f"%Sales%")
    ).scalar() or 0
    calls_service = call_query.with_entities(func.sum(QueueStatistics.calls)).filter(
        QueueStatistics.date.between(start_date, end_date),
        QueueStatistics.queue_name.like(f"%Service%")
    ).scalar() or 0

    # Email metrics
    total_emails = all_email_query.with_entities(func.sum(EmailData.received)).filter(
        EmailData.date.between(start_date, end_date)
    ).scalar() or 0

    sales_emails = email_query.with_entities(func.sum(WorkflowReportGuruKF.received)).filter(
        WorkflowReportGuruKF.date.between(start_date, end_date),
        WorkflowReportGuruKF.customer.notlike(f"%Service%")
    ).scalar() or 0

    service_emails = email_query.with_entities(func.sum(WorkflowReportGuruKF.received)).filter(
        WorkflowReportGuruKF.date.between(start_date, end_date),
        WorkflowReportGuruKF.customer.like(f"%Service%")
    ).scalar() or 0

    # Booking metrics 
    total_bookings = booking_query.with_entities(func.count(SoftBookingKF.id)).filter(
        SoftBookingKF.service_creation_time.between(start_date_booking, end_date_booking)
    ).scalar() or 0
    sb_bookings = booking_query.with_entities(func.count(SoftBookingKF.id)).filter(
        (SoftBookingKF.status == "OK") | (SoftBookingKF.status == "RF"), 
        SoftBookingKF.service_creation_time.between(start_date_booking, end_date_booking)
    ).scalar() or 0
    # print("Bookings: ", total_bookings, sb_bookings)

    # Apply domain filtering if domain is not "all"
    if domain != "all":
        if "Sales" in domain:
            call_query = call_query.filter(QueueStatistics.queue_name.notlike("%Service%"))
            all_call_query = all_call_query.filter(AllQueueStatisticsData.customer.notlike("%Service%"))
            email_query = email_query.filter(WorkflowReportGuruKF.customer.notlike("%Service%"))
            all_email_query = all_email_query.filter(EmailData.customer.notlike("%Service%"))
            booking_query = booking_query.filter(SoftBookingKF.customer.notlike("%Service%"))
        else:
            call_query = call_query.filter(QueueStatistics.queue_name.like(f"%{domain}%"))
            all_call_query = all_call_query.filter(AllQueueStatisticsData.customer.like(f"%{domain}%"))
            email_query = email_query.filter(WorkflowReportGuruKF.customer.like(f"%{domain}%"))
            all_email_query = all_email_query.filter(EmailData.customer.like(f"%{domain}%"))
            booking_query = booking_query.filter(SoftBookingKF.customer.like(f"%{domain}%"))

    # Daily calls trend
    calls_data = all_call_query.with_entities(
        AllQueueStatisticsData.date.label("date"),
        func.sum(AllQueueStatisticsData.calls).label("total_calls")
    ).filter(
        AllQueueStatisticsData.date.between(start_date, end_date)
    ).group_by(AllQueueStatisticsData.date).all()

    calls_trend = {
        row.date.strftime("%Y-%m-%d") if row.date else None: row.total_calls
        for row in calls_data
    }
    
    # Daily emails trend
    emails_data = all_email_query.with_entities(
        EmailData.date.label("date"), 
        func.sum(EmailData.received).label("total_emails")
    ).filter(
        EmailData.date.between(start_date, end_date)
    ).group_by(EmailData.date).all()
    email_trend = {
        row.date.strftime("%Y-%m-%d") if row.date else None: row.total_emails
        for row in emails_data
    }
    # print("calls trend: ", calls_trend)
    # print("emails trend: ", email_trend)
    # Other call metrics 
    calls_offered_data = all_call_query.with_entities(
        AllQueueStatisticsData.date.label("date"),
        func.sum(AllQueueStatisticsData.offered).label("offered_calls")
    ).filter(
        AllQueueStatisticsData.date.between(start_date, end_date)
    ).group_by(AllQueueStatisticsData.date).all()

    calls_offered = {
        row.date.strftime("%Y-%m-%d") if row.date else None: row.offered_calls
        for row in calls_offered_data
    }
    # print("calls offered :", calls_offered)
    # calls_offered = all_call_query.with_entities(func.sum(AllQueueStatisticsData.offered)).filter(
    #     AllQueueStatisticsData.date.between(yesterday_date, yesterday_date)
    # ).scalar() or 0
    calls_handled_data = all_call_query.with_entities(
        AllQueueStatisticsData.date.label("date"),
        func.sum(AllQueueStatisticsData.accepted).label("calls_handled")
    ).filter(
        AllQueueStatisticsData.date.between(start_date, end_date)
    ).group_by(AllQueueStatisticsData.date).all()

    calls_handled = {
        row.date.strftime("%Y-%m-%d") if row.date else None: row.calls_handled
        for row in calls_handled_data
    }
    # calls_handled = all_call_query.with_entities(func.sum(AllQueueStatisticsData.accepted)).filter(
    #     AllQueueStatisticsData.date.between(yesterday_date, yesterday_date)
    # ).scalar() or 0
    acc_data = all_call_query.with_entities(
        AllQueueStatisticsData.date.label("date"),
        func.sum(AllQueueStatisticsData.asr).label("acc")
    ).filter(
        AllQueueStatisticsData.date.between(start_date, end_date)
    ).group_by(AllQueueStatisticsData.date).all()

    acc = {
        row.date.strftime("%Y-%m-%d") if row.date else None: row.acc
        for row in acc_data
    }
    # acc = all_call_query.with_entities(func.avg(AllQueueStatisticsData.asr)).filter(
    #     AllQueueStatisticsData.date.between(yesterday_date, yesterday_date)
    # ).scalar() or 0
    sla_data = all_call_query.with_entities(
        AllQueueStatisticsData.date.label("date"),
        func.sum(AllQueueStatisticsData.sla_20_20).label("sla")
    ).filter(
        AllQueueStatisticsData.date.between(start_date, end_date)
    ).group_by(AllQueueStatisticsData.date).all()

    sla = {
        row.date.strftime("%Y-%m-%d") if row.date else None: row.sla
        for row in sla_data
    }
    # sla = all_call_query.with_entities(func.avg(AllQueueStatisticsData.sla_20_20)).filter(
    #     AllQueueStatisticsData.date.between(yesterday_date, yesterday_date)
    # ).scalar() or 0
    max_wait_time_data = all_call_query.with_entities(
        AllQueueStatisticsData.date.label("date"),
        func.sum(AllQueueStatisticsData.max_wait_time).label("max_wait_time")
    ).filter(
        AllQueueStatisticsData.date.between(start_date, end_date)
    ).group_by(AllQueueStatisticsData.date).all()

    max_wait_time = {
        row.date.strftime("%Y-%m-%d") if row.date else None:
        f"00:{str(int((row.max_wait_time or 0) / 60)).zfill(2)}:{str(int((row.max_wait_time or 0) % 60)).zfill(2)}"
        for row in max_wait_time_data
    }
    # print("Max time", max_wait_time_data, max_wait_time)
    # max_wait_time = all_call_query.with_entities(func.sum(AllQueueStatisticsData.max_wait_time)).filter(
    #     AllQueueStatisticsData.date.between(yesterday_date, yesterday_date)
    # ).scalar() or 0
    aht_data = all_call_query.with_entities(
        AllQueueStatisticsData.date.label("date"),
        func.sum(AllQueueStatisticsData.avg_handling_time_inbound).label("aht")
    ).filter(
        AllQueueStatisticsData.date.between(start_date, end_date)
    ).group_by(AllQueueStatisticsData.date).all()

    aht = {
        row.date.strftime("%Y-%m-%d") if row.date else None:
        f"00:{str(int((row.aht or 0) / 60)).zfill(2)}:{str(int((row.aht or 0) % 60)).zfill(2)}"
        for row in aht_data
    }
    # aht = all_call_query.with_entities(func.sum(AllQueueStatisticsData.avg_handling_time_inbound)).filter(
    #     AllQueueStatisticsData.date.between(yesterday_date, yesterday_date)
    # ).scalar() or 0
    total_call_time_data = all_call_query.with_entities(
        AllQueueStatisticsData.date.label("date"),
        func.sum(AllQueueStatisticsData.total_outbound_talk_time_destination).label("total_call_time")
    ).filter(
        AllQueueStatisticsData.date.between(start_date, end_date)
    ).group_by(AllQueueStatisticsData.date).all()

    total_call_time = {
        row.date.strftime("%Y-%m-%d") if row.date else None: row.total_call_time
        for row in total_call_time_data
    }
    # total_call_time = all_call_query.with_entities(func.sum(AllQueueStatisticsData.total_outbound_talk_time_destination)).filter(
    #     AllQueueStatisticsData.date.between(yesterday_date, yesterday_date)
    # ).scalar() or 0  
    call_outbound_data = all_call_query.with_entities(
        AllQueueStatisticsData.date.label("date"),
        func.sum(AllQueueStatisticsData.outbound_accepted).label("call_outbound")
    ).filter(
        AllQueueStatisticsData.date.between(start_date, end_date)
    ).group_by(AllQueueStatisticsData.date).all()

    call_outbound = {
        row.date.strftime("%Y-%m-%d") if row.date else None: row.call_outbound
        for row in call_outbound_data
    }
    # print("call outbound: ", call_outbound)
    # call_outbound = all_call_query.with_entities(func.sum(AllQueueStatisticsData.outbound_accepted)).filter(
    #     AllQueueStatisticsData.date.between(yesterday_date, yesterday_date)
    # ).scalar() or 0

    # Email processing metrics
    total_processing_time_seconds = 0.00001
    total_processing_time_min = 0
    total_processing_time_hour = 0
    service_level_gross = all_email_query.with_entities(func.avg(EmailData.service_level_gross)).filter(
        EmailData.date.between(yesterday_date, yesterday_date)
    ).scalar() or 0

    processing_times = all_email_query.with_entities(EmailData.processing_time).filter(
        EmailData.date.between(yesterday_date, yesterday_date)
    ).all()
    processing_times = [pt[0] if isinstance(pt, tuple) else pt for pt in processing_times]
    for pt in processing_times:
        hours, minutes, seconds = time_format(pt)
        total_processing_time_seconds += seconds
        total_processing_time_min += minutes
        total_processing_time_hour += hours

    total_processing_time_min += total_processing_time_seconds // 60
    total_processing_time_seconds = total_processing_time_seconds % 60
    total_processing_time_hour += total_processing_time_min // 60
    total_processing_time_min = total_processing_time_min % 60

    recieved_data = all_email_query.with_entities(
        EmailData.date.label("date"), 
        func.sum(EmailData.received).label("recieved")
    ).filter( 
        EmailData.date.between(start_date, end_date)
    ).group_by(EmailData.date).all()
    recieved = {
        row.date.strftime("%Y-%m-%d") if row.date else None: row.recieved
        for row in recieved_data}
    # print("Received: ", recieved)
    # recieved = all_email_query.with_entities(func.sum(EmailData.received)).filter(
    #     EmailData.date.between(yesterday_date, yesterday_date)
    # ).scalar() or 0
    sent_data = all_email_query.with_entities(
        EmailData.date.label("date"), 
        func.sum(EmailData.sent).label("sent")
    ).filter(
        EmailData.date.between(start_date, end_date)
    ).group_by(EmailData.date).all()
    sent = {
        row.date.strftime("%Y-%m-%d") if row.date else None: row.sent
        for row in sent_data}
    # sent = all_email_query.with_entities(func.sum(EmailData.sent)).filter(
    #     EmailData.date.between(yesterday_date, yesterday_date)
    # ).scalar() or 0
    archived_data = all_email_query.with_entities(
        EmailData.date.label("date"), 
        func.sum(EmailData.archived).label("archived")
    ).filter(
        EmailData.date.between(start_date, end_date)
    ).group_by(EmailData.date).all()
    archived = {
        row.date.strftime("%Y-%m-%d") if row.date else None: row.archived
        for row in archived_data}
    # archived = all_email_query.with_entities(func.sum(EmailData.archived)).filter(
    #     EmailData.date.between(yesterday_date, yesterday_date)
    # ).scalar() or 0 
    all_emails_data = email_query.with_entities(
        WorkflowReportGuruKF.date.label("date"), 
        func.count(WorkflowReportGuruKF.id).label("all_emails")
    ).filter(
        WorkflowReportGuruKF.date.between(start_date, end_date)
    ).group_by(WorkflowReportGuruKF.date).all()
    all_emails = {
        row.date.strftime("%Y-%m-%d") if row.date else None: row.all_emails
        for row in all_emails_data}
    # all_emails = email_query.with_entities(func.count(WorkflowReportGuruKF.id)).filter(
    #     WorkflowReportGuruKF.date.between(yesterday_date, yesterday_date)
    # ).scalar() or 0

    # booking metrics 
    all_bookings_data = booking_query.with_entities(
        func.date(SoftBookingKF.service_creation_time).label("date"),
        func.count(SoftBookingKF.id).label("all_bookings")
    ).filter(
        SoftBookingKF.service_creation_time.between(start_date_booking, end_date_booking)
    ).group_by(func.date(SoftBookingKF.service_creation_time)).all()

    all_bookings = {
        str(row.date) if isinstance(row.date, int) else (
            row.date.strftime("%Y-%m-%d") if isinstance(row.date, datetime) else str(row.date)
        ): row.all_bookings
        for row in all_bookings_data
    }
    # all_bookings_data = booking_query.with_entities(
    #     func.date(SoftBookingKF.service_creation_time).label("date"),
    #     func.count(SoftBookingKF.id).label("all_bookings")
    # ).filter(
    #     SoftBookingKF.service_creation_time.between(start_date_booking, end_date_booking)
    # ).group_by(
    #     func.date(SoftBookingKF.service_creation_time)
    # ).all()

    # all_bookings = {
    #     row.date: row.all_bookings
    #     for row in all_bookings_data
    # }
    
    # print("Bookings: ", all_bookings)
        # yesterday_bookings = booking_query.with_entities(func.count(SoftBookingKF.id)).filter(
    #     SoftBookingKF.service_creation_time.between(yesterday_date_start, yesterday_date_end)
    # ).scalar() or 0
    booked_data = booking_query.with_entities(
        func.date(SoftBookingKF.service_creation_time).label("date"),
        func.count(SoftBookingKF.id).label("booked")
    ).filter(
        (SoftBookingKF.status == "OK") | (SoftBookingKF.status == "RF"), 
        SoftBookingKF.service_creation_time.between(start_date_booking, end_date_booking)
    ).group_by(func.date(SoftBookingKF.service_creation_time)).all() 
    booked = {
        row.date.strftime("%Y-%m-%d") if isinstance(row.date, datetime) else str(row.date): row.booked
        for row in booked_data
    }
    # print("Booked :", booked)
    # booked = booking_query.with_entities(func.count(SoftBookingKF.id)).filter(
    #     (SoftBookingKF.status == "OK") | (SoftBookingKF.status == "RF"), 
    #     SoftBookingKF.service_creation_time.between(yesterday_date_start, yesterday_date_end)
    # ).scalar() or 0
    not_booked_data = booking_query.with_entities(
        func.date(SoftBookingKF.service_creation_time).label("date"),
        func.count(SoftBookingKF.id).label("not_booked")
    ).filter(
        (SoftBookingKF.status == "XX"), 
        SoftBookingKF.service_creation_time.between(start_date_booking, end_date_booking)
    ).group_by(func.date(SoftBookingKF.service_creation_time)).all() 
    not_booked = {
        row.date.strftime("%Y-%m-%d") if isinstance(row.date, datetime) else str(row.date): row.not_booked
        for row in not_booked_data
    }
    # not_booked = booking_query.with_entities(func.count(SoftBookingKF.id)).filter(
    #     (SoftBookingKF.status == "XX"), 
    #     SoftBookingKF.service_creation_time.between(yesterday_date_start, yesterday_date_end)
    # ).scalar() or 0
    pending_data = booking_query.with_entities(
        func.date(SoftBookingKF.service_creation_time).label("date"),
        func.count(SoftBookingKF.id).label("pending")
    ).filter(
        (SoftBookingKF.status == "PE"), 
        SoftBookingKF.service_creation_time.between(start_date_booking, end_date_booking)
    ).group_by(func.date(SoftBookingKF.service_creation_time)).all() 
    pending = {
        row.date.strftime("%Y-%m-%d") if isinstance(row.date, datetime) else str(row.date): row.pending
        for row in pending_data
    }
    # pending = booking_query.with_entities(func.count(SoftBookingKF.id)).filter(
    #     (SoftBookingKF.status == "PE"), 
    #     SoftBookingKF.service_creation_time.between(yesterday_date_start, yesterday_date_end)
    # ).scalar() or 0
    op_rq_data = booking_query.with_entities(
        func.date(SoftBookingKF.service_creation_time).label("date"),
        func.count(SoftBookingKF.id).label("op_rq")
    ).filter(
        (SoftBookingKF.status == "OP") | (SoftBookingKF.status == "RQ"), 
        SoftBookingKF.service_creation_time.between(start_date_booking, end_date_booking)
    ).group_by(func.date(SoftBookingKF.service_creation_time)).all() 
    op_rq = {
        row.date.strftime("%Y-%m-%d") if isinstance(row.date, datetime) else str(row.date): row.op_rq
        for row in op_rq_data
    }
    # op_rq = booking_query.with_entities(func.count(SoftBookingKF.id)).filter(
    #     (SoftBookingKF.status == "OP") | (SoftBookingKF.status == "RQ"), 
    #     SoftBookingKF.service_creation_time.between(yesterday_date_start, yesterday_date_end)
    # ).scalar() or 0
    sb_data = booking_query.with_entities(
        func.date(SoftBookingKF.service_creation_time).label("date"),
        func.count(SoftBookingKF.id).label("sb")
    ).filter(
        (SoftBookingKF.status == "SB"), 
        SoftBookingKF.service_creation_time.between(start_date_booking, end_date_booking)
    ).group_by(func.date(SoftBookingKF.service_creation_time)).all() 
    sb = {
        row.date.strftime("%Y-%m-%d") if isinstance(row.date, datetime) else str(row.date): row.sb
        for row in sb_data
    }
    # sb = booking_query.with_entities(func.count(SoftBookingKF.id)).filter(
    #     (SoftBookingKF.status == "SB"), 
    #     SoftBookingKF.service_creation_time.between(yesterday_date_start, yesterday_date_end)
    # ).scalar() or 0
    total_sb_booked_and_cancelled = {
        date: booked.get(date, 0) + sb.get(date, 0) + not_booked.get(date, 0)
        for date in set(booked) | set(sb) | set(not_booked)
    }

    booking_rate = {
        date: round((booked.get(date, 0) + sb.get(date, 0)) / total * 100, 2) if total > 0 else 0
        for date, total in total_sb_booked_and_cancelled.items()
    }
    # print("Booking rate: ", booking_rate)
    # total_sb_booked_and_cancelled = booked + sb + not_booked
    # booking_rate = (sb + booked / total_sb_booked_and_cancelled * 100) if total_sb_booked_and_cancelled > 0 else 0
  
    return {
        "total_bookings": total_bookings,
        "calls": calls,
        "calls_sales": calls_sales,
        "calls_service": calls_service,
        "total_emails": total_emails,
        "mails_sales": sales_emails, 
        "mails_service": service_emails,
        "sb_bookings": sb_bookings,
        "table": {
            "calls": calls_trend,
            "emails": email_trend,
        },
        "call_metrics": {
            "calls_offered": calls_offered,
            "calls_handled": calls_handled,
            "ASR": acc,
            "SLA": sla,
            "Max Waiting Time": max_wait_time,
            "AHT": aht,
            "Total_Call_Time": total_call_time,
            "Outbound_Calls": call_outbound,
            "sales_total_calls": all_calls,
            "guru_wrong_calls": guru_wrong_call,
            "guru_bookings": guru_bookings,
            "true_sales_calls_guru": true_guru_calls,
            "organisch_conversion":organisch_conversion,
            "cb_wrong_calls":cb_wrong_call,
            "cb_bookings": cb_bookings,
            "true_sales_calls_cb": true_cb_calls,
            "cb_conversion": cb_conversion,
        },
        "email_metrics": {
            "all_emails": all_emails,
            "received": recieved,
            "sent": sent,
            "archived": archived,
            "Total_Processing_Time": time_formatter(int(total_processing_time_hour), int(total_processing_time_min), int(total_processing_time_seconds)),
        },
        "booking_metrics": {
            "bookings": all_bookings,
            "booked": booked,
            "not_booked": not_booked,
            "bookings_pending": pending,
            "bookings_op_rq": op_rq,
            "bookings_sb": sb,
            "booking_rate": booking_rate
        },
    } 


def create_header_style():
    return Font(bold=True), PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')


def apply_cell_border(cell):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def create_section(workbook, start_row, section_name, dates, data, bold_font, section_type="Sales"):
    section_height = 6  # Table height for calls and mails
    print("started writing data")
    workbook[f'A{start_row}'] = section_name
    workbook.merge_cells(f'A{start_row}:A{start_row + section_height - 1}')
    workbook[f'A{start_row}'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    workbook[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
    workbook[f'A{start_row}'].font = bold_font

    calls_row = start_row + 3
    mails_row = start_row + 5
    workbook[f'B{calls_row}'] = "Calls"
    workbook[f'B{mails_row}'] = "Mails"
    workbook[f'B{calls_row}'].font = bold_font
    workbook[f'B{mails_row}'].font = bold_font

    days = ['Samstag', 'Sonntag', 'Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag']
    col_offset = 2  # Start from column C
    calls_data = data['table']['calls']
    emails_data = data['table']['emails']

    for idx, date in enumerate(dates):
        col = idx + col_offset + 1
        date_str = date.strftime('%Y-%m-%d')

        workbook.cell(row=start_row, column=col, value=days[date.weekday()])
        workbook.cell(row=start_row + 1, column=col, value=date.strftime('%Y/%m/%d'))

        call_value = calls_data.get(date_str, 0)
        email_value = emails_data.get(date_str, 0)

        if section_type == "Service":
            call_value = int(call_value * 0.857)

        workbook.cell(row=calls_row, column=col, value=call_value)
        workbook.cell(row=mails_row, column=col, value=email_value)

    for row in range(start_row, start_row + section_height):
        for col in range(1, len(dates) + col_offset + 1):
            cell = workbook.cell(row=row, column=col)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                apply_cell_border(cell)
                if row < start_row + 3:
                    cell.alignment = Alignment(horizontal='center')

    next_row = start_row + section_height + 2  # Space before metrics

    # ---- Call Metrics ----
    workbook.cell(row=next_row, column=1, value="Call Metrics").font = bold_font
    next_row += 1
    call_metrics = [
        ('Calls offered', 'calls_offered'),
        ('Calls handled', 'calls_handled'),
        ('ASR', 'ASR'),
        ('SLA', 'SLA'),
        ('Längste Wartezeit', 'Max Waiting Time'),
        ('AHT', 'AHT'),
        ('Gesprächszeit gesamt', 'Total_Call_Time'),
        ('Call Outbound', 'Outbound_Calls')
    ]
    
    for label, field in call_metrics:
        workbook.cell(row=next_row, column=1, value=label) 
        for idx, date in enumerate(dates):
            col = idx + col_offset + 1
            date_str = date.strftime('%Y-%m-%d')
            workbook.cell(row=next_row, column=col, value=data['call_metrics'].get(field, {}).get(date_str, 0))
        next_row += 1

    next_row += 2 

    # ---- Email Metrics ----
    workbook.cell(row=next_row, column=1, value="Email Metrics").font = bold_font
    next_row += 1
    email_metrics = [
        ('Mails', 'all_emails'),
        ('Empfangen', 'received'),
        ('Gesendet', 'sent'),
        ('Archiviert', 'archived'),
        # ('Bearbeitungszeit gesamt', 'Total_Processing_Time')
    ]
    
    for label, field in email_metrics:
        workbook.cell(row=next_row, column=1, value=label)
        for idx, date in enumerate(dates):
            col = idx + col_offset + 1
            date_str = date.strftime('%Y-%m-%d')
            workbook.cell(row=next_row, column=col, value=data['email_metrics'].get(field, {}).get(date_str, 0))
        next_row += 1

    # next_row += 2  # Space before booking metrics

    # # ---- Booking Metrics ----
    # workbook.cell(row=next_row, column=1, value="Buchungen Metrics").font = bold_font
    # next_row += 1
    # booking_metrics = [
    #     ('Gebucht', 'booked'),
    #     ('Nicht gebucht', 'not_booked'),
    #     ('Pending', 'bookings_pending'),
    #     ('OP/RQ', 'bookings_op_rq'),
    #     ('SB', 'bookings_sb'),
    #     ('SB Buchungsquote', 'booking_rate'),
    #     ('Buchungen gesamt', 'bookings'),
    # ]
    
    # for label, field in booking_metrics:
    #     workbook.cell(row=next_row, column=1, value=label)
    #     for idx, date in enumerate(dates):
    #         col = idx + col_offset + 1
    #         date_str = date.strftime('%Y-%m-%d')
    #         print(data['booking_metrics'].get(field, {}).get(date_str, 0))
    #         workbook.cell(row=next_row, column=col, value=data['booking_metrics'].get(field, {}).get(date_str, 0))
    #     next_row += 1

    return next_row + 2 

def write_sheet_data(ws, result, sheet_prefix, start_date, end_date, split_domains=False, result_service=None):
    """
    Writes structured data in the worksheet in this format:
    
    - Sales Table (if split)
    - Sales Metrics (if split)
    - Service Table (if split)
    - Service Metrics (if split)
    - Single Table (if no split)
    """
    bold_font = Font(bold=True)
    ws['A1'] = f"{sheet_prefix} {datetime.now().date()}"
    ws['A2'] = str(datetime.now().year)
    ws['A2'].font = bold_font

    # Write the main metrics header
    headers = [
        ('B2', 'Buchungen'),
        ('C2', 'Kontakte'),
        ('D2', 'Call'),
        ('E2', 'Mail'),
        ('F2', 'Mail Sales'),
        ('G2', 'Mail Service'),
        ('H2', 'Call Sales'),
        ('I2', 'Call Service'),
        ('J2', 'SB')
    ]
    for cell, value in headers:
        ws[cell] = value
        ws[cell].font = bold_font

    # Populate the main metrics row
    if split_domains and result_service:
        ws['B3'] = result['total_bookings'] + result_service['total_bookings']
        ws['C3'] = (result['calls'] + result['total_emails']) + (result_service['calls'] + result_service['total_emails'])
        ws['D3'] = result['calls'] + result_service['calls']
        ws['E3'] = result['total_emails'] + result_service['total_emails']
        ws['F3'] = result['mails_sales']
        ws['G3'] = result_service['mails_service']
        ws['H3'] = result['calls_sales']
        ws['I3'] = result_service['calls_service']
        ws['J3'] = result['sb_bookings'] + result_service['sb_bookings']
    else:
        ws['B3'] = result['total_bookings']
        ws['C3'] = result['calls'] + result['total_emails']
        ws['D3'] = result['calls']
        ws['E3'] = result['total_emails']
        ws['F3'] = result['mails_sales']
        ws['G3'] = result['mails_service']
        ws['H3'] = result['calls_sales']
        ws['I3'] = result['calls_service']
        ws['J3'] = result['sb_bookings']

    # Generate the date range from `start_date` to `end_date`
    dates = []
    current_date = start_date
    while current_date <= end_date:
        dates.append(current_date)
        current_date += timedelta(days=1)

    next_row = 5  # Start after the main metrics row

    # ---- Sales Section ----
    if split_domains:
        ws.cell(row=next_row, column=1, value=f"{sheet_prefix} Sales").font = bold_font
        next_row += 1
        next_row = create_section(ws, next_row, f"{sheet_prefix} Sales", dates, result, bold_font)
        next_row += 2  # Space before metrics

        # Add Sales Metrics
        # ws.cell(row=next_row, column=1, value="Sales Metrics").font = bold_font
        # next_row += 1
        # next_row = write_metrics(ws, next_row, result, bold_font)

    # ---- Service Section ----
    if split_domains and result_service:
        next_row += 2  # Add some space
        ws.cell(row=next_row, column=1, value=f"{sheet_prefix} Service").font = bold_font
        next_row += 1
        next_row = create_section(ws, next_row, f"{sheet_prefix} Service", dates, result_service, bold_font)
        next_row += 2  # Space before metrics
        
        # ---- Conversion Metrics ----
        ws.cell(row=next_row, column=1, value="Conversion Metrics").font = bold_font
        next_row += 1
        conversion_metrics = [
            ('Gesamtanzahl der Anrufe', 'sales_total_calls'),
            ('Guru Wrong Calls', 'guru_wrong_calls'),
            ('echter Verkaufsgesprächsguru', 'true_sales_calls_guru'),
            ('Guru-Buchungen', 'guru_bookings'),
            ('Verkaufskonversion', 'organisch_conversion'),
            ('cb falsche Anrufe', 'cb_wrong_calls'),
            ('CB-Buchungen', 'cb_bookings'),
            ('echte Verkaufsgespräche cb', 'true_sales_calls_cb'),
            ('CB konversion', 'cb_conversion'),
        ]
        
        # for label, field in conversion_metrics:
        #     ws.cell(row=next_row, column=1, value=label)
        #     for idx, date in enumerate(dates):
        #         col = idx + 2 + 1
        #         date_str = date.strftime('%Y-%m-%d')
        #         # print(data['booking_metrics'].get(field, {}).get(date_str, 0))
        #         ws.cell(row=next_row, column=col, value=result['call_metrics'].get(field, {}).get(date_str, 0))
        #     next_row += 1
        # For the conversion metrics section
        for label, field in conversion_metrics:
            ws.cell(row=next_row, column=1, value=label)
            for idx, date in enumerate(dates):
                col = idx + 2 + 1
                date_str = date.strftime('%Y-%m-%d')
                # Check if the field exists and is a dictionary before trying to access date_str
                field_data = result['call_metrics'].get(field, {})
                if isinstance(field_data, dict):
                    value = field_data.get(date_str, 0)
                else:
                    value = field_data  # Use the value directly if it's not a dict
                ws.cell(row=next_row, column=col, value=value)
            next_row += 1
        next_row += 2
        # ---- Booking Metrics ----
        ws.cell(row=next_row, column=1, value="Buchungen Metrics").font = bold_font
        next_row += 1
        booking_metrics = [
            ('Gebucht', 'booked'),
            ('Nicht gebucht', 'not_booked'),
            ('Pending', 'bookings_pending'),
            ('OP/RQ', 'bookings_op_rq'),
            ('SB', 'bookings_sb'),
            ('SB Buchungsquote', 'booking_rate'),
            ('Buchungen gesamt', 'bookings'),
        ]
        
        for label, field in booking_metrics:
            ws.cell(row=next_row, column=1, value=label)
            for idx, date in enumerate(dates):
                col = idx + 2 + 1
                date_str = date.strftime('%Y-%m-%d')
                # print(data['booking_metrics'].get(field, {}).get(date_str, 0))
                ws.cell(row=next_row, column=col, value=result['booking_metrics'].get(field, {}).get(date_str, 0))
            next_row += 1

        # Add Service Metrics
        # ws.cell(row=next_row, column=1, value="Service Metrics").font = bold_font
        # next_row += 1
        # next_row = write_metrics(ws, next_row, result_service, bold_font)
    
    # If no splitting is needed, write everything as a single block.
    if not split_domains:
        ws.cell(row=next_row, column=1, value=f"{sheet_prefix} Overview").font = bold_font
        next_row += 1
        next_row = create_section(ws, next_row, f"{sheet_prefix}", dates, result, bold_font)
        next_row += 2  # Space before metrics
        # ---- Booking Metrics ----
        ws.cell(row=next_row, column=1, value="Buchungen Metrics").font = bold_font
        next_row += 1
        booking_metrics = [
            ('Gebucht', 'booked'),
            ('Nicht gebucht', 'not_booked'),
            ('Pending', 'bookings_pending'),
            ('OP/RQ', 'bookings_op_rq'),
            ('SB', 'bookings_sb'),
            ('SB Buchungsquote', 'booking_rate'),
            ('Buchungen gesamt', 'bookings'),
        ]
        
        for label, field in booking_metrics:
            ws.cell(row=next_row, column=1, value=label)
            for idx, date in enumerate(dates):
                col = idx + 2 + 1
                date_str = date.strftime('%Y-%m-%d')
                # print(data['booking_metrics'].get(field, {}).get(date_str, 0))
                ws.cell(row=next_row, column=col, value=result['booking_metrics'].get(field, {}).get(date_str, 0))
            next_row += 1
        # Add Metrics
        # ws.cell(row=next_row, column=1, value="Metrics").font = bold_font
        # next_row += 1
        # next_row = write_metrics(ws, next_row, result, bold_font)

    next_row+=2
    # Add Service Metrics
    # ws.cell(row=next_row, column=1, value="Buchugen Metrics").font = bold_font
    # next_row += 1
    # next_row = write_booking_metrics(ws, next_row, result, bold_font)
    
    return next_row + len(headers)  


@router.post("/export/excel")
async def export_excel(
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(oauth2.get_current_user),
    month: Optional[str] = Query(None, description="Month in YYYY-MM format (optional)"),
):
    user = db.query(User).filter(User.email == current_user.get("email")).first()
    
    # Get permission-based date range (with optional month override)
    start_date, end_date = validate_user_and_date_permissions_export(db=db, current_user=current_user, month=month)
    # start_date_booking, end_date_booking = validate_user_and_date_permissions_booking_export(db=db, current_user=current_user)
    start_date_booking = datetime.combine(start_date, datetime.min.time())
    end_date_booking = datetime.combine(end_date, datetime.max.time())
    
    # print("Final Date Range:", start_date, "to", end_date)

    is_admin_or_employee = user.role in ["admin", "employee"]
    if is_admin_or_employee:
        companies_to_export = ["5vorflug", "bild", "guru", "Galeria", "ADAC", "Urlaub"]
    else:
        accessible_companies, _, _ = domains_checker(db, user.id, filter_5vf="5vorFlug", filter_bild="BILD")
        companies_to_export = accessible_companies

    split_companies = ["5vorflug", "guru", "bild"]

    output = BytesIO()
    wb = openpyxl.Workbook()
    first_sheet = True

    for company in companies_to_export:
        acc_companies, call_filters, all_call_filters = domains_checker(
            db, user.id, filter_5vf="5vorFlug", filter_bild="BILD", target_company=company
        )
        acc_companies_email, email_filters, all_email_filters = domains_checker_email(
            db, user.id, filter_5vf="5vorFlug", filter_bild="Bild", target_company=company
        )
        acc_companies_booking, booking_filters, order_filters = domains_checker_booking(
            db, user.id, filter_5vf="5vF", filter_bild="BILD", target_company=company
        )

        if company.lower() not in [c.lower() for c in acc_companies]:
            continue

        call_query = db.query(QueueStatistics).filter(or_(*call_filters)) if call_filters else db.query(QueueStatistics)
        all_call_query = db.query(AllQueueStatisticsData).filter(or_(*all_call_filters)) if all_call_filters else db.query(AllQueueStatisticsData)
        email_query = db.query(WorkflowReportGuruKF).filter(or_(*email_filters)) if email_filters else db.query(WorkflowReportGuruKF)
        all_email_query = db.query(EmailData).filter(or_(*all_email_filters)) if all_email_filters else db.query(EmailData)
        booking_query = db.query(SoftBookingKF).filter(or_(*booking_filters)) if booking_filters else db.query(SoftBookingKF)

        if first_sheet:
            ws = wb.active
            ws.title = company
            first_sheet = False
        else:
            ws = wb.create_sheet(title=company)

        if company.lower() in split_companies:
            result_sales = get_export_data(
                call_query=call_query,
                all_call_query=all_call_query,
                email_query=email_query,
                all_email_query=all_email_query,
                booking_query=booking_query,
                start_date=start_date,
                end_date=end_date,
                start_date_booking=start_date_booking,
                end_date_booking=end_date_booking,
                domain="Sales",
                company=company
            )
            result_service = get_export_data(
                call_query=call_query,
                all_call_query=all_call_query,
                email_query=email_query,
                all_email_query=all_email_query,
                booking_query=booking_query,
                start_date=start_date,
                end_date=end_date,
                start_date_booking=start_date_booking,
                end_date_booking=end_date_booking,
                domain="Service",
                company=company
            )
            write_sheet_data(ws, result_sales, company, start_date, end_date, split_domains=True, result_service=result_service)
        else:
            result = get_export_data(
                call_query=call_query,
                all_call_query=all_call_query,
                email_query=email_query,
                all_email_query=all_email_query,
                booking_query=booking_query,
                start_date=start_date,
                end_date=end_date,
                start_date_booking=start_date_booking,
                end_date_booking=end_date_booking,
                domain="Sales",
                company=company
            )
            write_sheet_data(ws, result, company, start_date, end_date, split_domains=False)


    wb.save(output)
    output.seek(0)
    headers = {
        'Content-Disposition': f'attachment; filename="FC_neu_{datetime.now().strftime("%d_%m_%Y")}.xlsx"'
    }
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers
    )

